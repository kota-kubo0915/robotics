import openpyxl
import re
from statistics import mean

tname = '6.txt'
bname = 'フリーライダー50個60人500step.csv'

file = open('6.txt')
lines = file.readlines()

wb = openpyxl.Workbook() #Book新規作成
#wb = openpyxl.load_workbook(bname)
#wb.create_sheet(title='12 500step')
ws = wb.active

ws.cell(1,1).value = lines[50].split(":")[0]
ws.cell(1,2).value = float(lines[50].split(":")[1])

ws.cell(2,1).value = lines[51].split(":")[0]
ws.cell(2,2).value = float(lines[51].split(":")[1])

ws.cell(3,1).value = lines[52].split(":")[0]
ws.cell(3,2).value = float(lines[52].split(":")[1])

ws.cell(4,1).value = lines[53].split(":")[0]
ws.cell(4,2).value = float(lines[53].split(":")[1])

for i in range(50):
    ws.cell(i+5,1).value = lines[i].split("\t")[0]

for j in range(100):
    for i in range(50):
        ws.cell(2*i+55+102*j,1).value = lines[2*i+54+102*j].split("\t")[0]
        ws.cell(2*i+56+102*j,1).value = lines[2*i+55+102*j].split("{")[0]
        ws.cell(2*i+56+102*j,2).value = lines[2*i+55+102*j].split("{")[1] and lines[2*i+55+102*j].split(",")[0]
        ws.cell(2*i+56+102*j,3).value = lines[2*i+55+102*j].split(",")[1]
        ws.cell(2*i+56+102*j,4).value = lines[2*i+55+102*j].split(",")[2]
        ws.cell(2*i+56+102*j,5).value = lines[2*i+55+102*j].split(",")[3] and lines[2*i+55+102*j].split("}")[0]
        ws.cell(2*i+55+102*j,6).value = float(lines[2*i+55+102*j].split("=")[1])
    j = j+1


for i in range(99):
    ws.cell(102*i+155,1).value = lines[102*i+154].split(":")[0]
    ws.cell(102*i+155,2).value = float(lines[102*i+154].split(":")[1])
    ws.cell(102*i+156,1).value = lines[102*i+155].split("\t")[0]

for j in range(100):
    for i in range(50):
        a = ws.cell(2*i+56+102*j,3).value
        ws.cell(2*i+56+102*j,3).value = float(a.replace("'s':",""))
        b = ws.cell(2*i+56+102*j,4).value
        ws.cell(2*i+56+102*j,4).value = float(b.replace("'l':",""))
        c = ws.cell(2*i+56+102*j,2).value
        ws.cell(2*i+56+102*j,2).value = float(c[:1] +''+ c[12:]) #桁数によって変える必要がある．
        d = ws.cell(2*i+56+102*j,5).value
        ws.cell(2*i+56+102*j,5).value = d[:1] +''+ d[37:] #桁数によって変える必要がある．
        print(len(c),c)
        print(len(d),d)

for j in range(100): #状況によっては加える
    for i in range(50):
        d = ws.cell(2*i+56+102*j,5).value
        ws.cell(2*i+56+102*j,5).value = d.replace(':','')

#ws.cell(row=8,column=2).value = lines[7].replace("Str =",'')

ws.cell(1,8).value = '協力的'
ws.cell(1,9).value = '懐疑的'
ws.cell(1,10).value = 'フリーライダー'
ws.cell(1,11).value = '嘘つき'
ws.cell(1,12).value = '回収個数'

for j in range(100):
    for i in range(50):
        ws.cell(i+2+51*j,8).value = ws.cell(2*i+56+102*j,2).value
        ws.cell(i+2+51*j,9).value = ws.cell(2*i+56+102*j,3).value
        ws.cell(i+2+51*j,10).value = ws.cell(2*i+56+102*j,4).value
        ws.cell(i+2+51*j,11).value = ws.cell(2*i+56+102*j,5).value
        ws.cell(i+2+51*j,12).value = ws.cell(2*i+55+102*j,6).value
    j = j+1

for i in range(100):
    ws.cell(51*i+52,8).value = '-----------------------'

for j in range(100):
    for i in range(50):
        ws.cell(j+2,i+15).value=ws.cell(i+2+51*j,8).value
        ws.cell(j+2,i+66).value=ws.cell(i+2+51*j,9).value
        ws.cell(j+2,i+117).value=ws.cell(i+2+51*j,10).value
        ws.cell(j+2,i+168).value=ws.cell(i+2+51*j,11).value
        ws.cell(j+2,i+219).value=ws.cell(i+2+51*j,12).value
    j = j+1

#for i in range(1,100):
    #ws.cell(i+1,15).value = '= AVERAGE(H:H11)'

#wb.save('繰り返しゲームga回収物.xlsx')
wb.save(bname)
