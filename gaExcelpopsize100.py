import openpyxl
import re
from statistics import mean

tname = 'l48.txt'
bname = '繰り返しゲームga 100.xlsx'

file = open('l48.txt')
lines = file.readlines()

#wb = openpyxl.Workbook() #Book新規作成
wb = openpyxl.load_workbook(bname)
wb.create_sheet(title='48 1000step')
ws = wb['48 1000step']


#初期設定の表示
ws.cell(1,1).value = lines[100].split(":")[0]
ws.cell(1,2).value = float(lines[100].split(":")[1])#Max Trial

ws.cell(2,1).value = lines[101].split(":")[0]
ws.cell(2,2).value = float(lines[101].split(":")[1])#Generation

ws.cell(3,1).value = lines[102].split(":")[0]
ws.cell(3,2).value = float(lines[102].split(":")[1])#pop_size

ws.cell(4,1).value = lines[103].split(":")[0]
ws.cell(4,2).value = float(lines[103].split(":")[1])#MaxStep

#ランダムに作成された遺伝子の表示
for i in range(100):
    ws.cell(i+5,1).value = lines[i].split("\t")[0]

for j in range(100):
    for i in range(100):
        ws.cell(2*i+105+202*j,1).value = lines[2*i+104+202*j].split("\t")[0]
        ws.cell(2*i+106+202*j,1).value = lines[2*i+105+202*j].split("{")[0]
        ws.cell(2*i+106+202*j,2).value = lines[2*i+105+202*j].split("{")[1] and lines[2*i+105+202*j].split(",")[0]
        ws.cell(2*i+106+202*j,3).value = lines[2*i+105+202*j].split(",")[1]
        ws.cell(2*i+106+202*j,4).value = lines[2*i+105+202*j].split(",")[2]
        ws.cell(2*i+106+202*j,5).value = lines[2*i+105+202*j].split(",")[3] and lines[2*i+105].split("}")[0]
        ws.cell(2*i+105+202*j,6).value = float(lines[2*i+105+202*j].split("=")[1])
    j = j+1


for i in range(99):
    ws.cell(202*i+305,1).value = lines[202*i+304].split(":")[0]
    ws.cell(202*i+305,2).value = float(lines[202*i+304].split(":")[1])
    ws.cell(202*i+306,1).value = lines[202*i+305].split("\t")[0]

for j in range(100):
    for i in range(100):
        a = ws.cell(2*i+106+202*j,3).value
        ws.cell(2*i+106+202*j,3).value = float(a.replace("'s':",""))
        b = ws.cell(2*i+106+202*j,4).value
        ws.cell(2*i+106+202*j,4).value = float(b.replace("'l':",""))
        c = ws.cell(2*i+106+202*j,2).value
        ws.cell(2*i+106+202*j,2).value = float(c[:1] +''+ c[12:]) #桁数によって変える必要がある．
        d = ws.cell(2*i+106+202*j,5).value
        ws.cell(2*i+106+202*j,5).value = d[:1] +''+ d[37:] #桁数によって変える必要がある．
        #print(len(c),c)
        #print(len(d),d)

for j in range(100): #状況によっては加える
    for i in range(100):
        d = ws.cell(2*i+106+202*j,5).value
        ws.cell(2*i+106+202*j,5).value = float(d.replace(':',''))

#ws.cell(row=8,column=2).value = lines[7].replace("Str =",'')

ws.cell(1,8).value = '協力的'
ws.cell(1,9).value = '懐疑的'
ws.cell(1,10).value = 'フリーライダー'
ws.cell(1,11).value = '嘘つき'
ws.cell(1,12).value = '回収個数'

for j in range(100):
    for i in range(100):
        ws.cell(i+2+101*j,8).value = ws.cell(2*i+106+202*j,2).value
        ws.cell(i+2+101*j,9).value = ws.cell(2*i+106+202*j,3).value
        ws.cell(i+2+101*j,10).value = ws.cell(2*i+106+202*j,4).value
        ws.cell(i+2+101*j,11).value = ws.cell(2*i+106+202*j,5).value
        ws.cell(i+2+101*j,12).value = ws.cell(2*i+105+202*j,6).value
    j = j+1

for i in range(100):
    ws.cell(101*i+102,8).value = '-----------------------'

for j in range(100):
    for i in range(100):
        ws.cell(j+2,i+15).value=ws.cell(i+2+101*j,8).value
        ws.cell(j+2,i+116).value=ws.cell(i+2+101*j,9).value
        ws.cell(j+2,i+217).value=ws.cell(i+2+101*j,10).value
        ws.cell(j+2,i+318).value=ws.cell(i+2+101*j,11).value
        ws.cell(j+2,i+419).value=ws.cell(i+2+101*j,12).value
    j = j+1

#for i in range(1,100):
    #ws.cell(i+1,15).value = '= AVERAGE(H:H11)'

#wb.save('繰り返しゲームga回収物.xlsx')
wb.save(bname)
