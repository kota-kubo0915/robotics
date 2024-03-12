import openpyxl
import re
from statistics import mean

tname = 'l.txt'
bname = '繰り返しゲームga.xlsx'

file = open('l.txt')
lines = file.readlines()

#wb = openpyxl.Workbook() #Book新規作成
wb = openpyxl.load_workbook(bname)
#wb.create_sheet(title='各12人 500step')
ws = wb['36 500step']

ws.cell(1,1).value = lines[10].split(":")[0]
ws.cell(1,2).value = float(lines[10].split(":")[1])

ws.cell(2,1).value = lines[11].split(":")[0]
ws.cell(2,2).value = float(lines[11].split(":")[1])

ws.cell(3,1).value = lines[12].split(":")[0]
ws.cell(3,2).value = float(lines[12].split(":")[1])

ws.cell(4,1).value = lines[13].split(":")[0]
ws.cell(4,2).value = float(lines[13].split(":")[1])

for i in range(10):
    ws.cell(i+5,1).value = lines[i].split("\t")[0]

for j in range(100):
    for i in range(10):
        ws.cell(2*i+15+22*j,1).value = lines[2*i+14+22*j].split("\t")[0]
        ws.cell(2*i+16+22*j,1).value = lines[2*i+15+22*j].split("{")[0]
        ws.cell(2*i+16+22*j,2).value = lines[2*i+15+22*j].split("{")[1] and lines[2*i+15+22*j].split(",")[0]
        ws.cell(2*i+16+22*j,3).value = lines[2*i+15+22*j].split(",")[1]
        ws.cell(2*i+16+22*j,4).value = lines[2*i+15+22*j].split(",")[2]
        ws.cell(2*i+16+22*j,5).value = lines[2*i+15+22*j].split(",")[3] and lines[2*i+15+22*j].split("}")[0]
        ws.cell(2*i+15+22*j,6).value = float(lines[2*i+15+22*j].split("=")[1])
    j = j+1


for i in range(99):
    ws.cell(22*i+35,1).value = lines[22*i+34].split(":")[0]
    ws.cell(22*i+35,2).value = float(lines[22*i+34].split(":")[1])
    ws.cell(22*i+36,1).value = lines[22*i+35].split("\t")[0]

for j in range(100):
    for i in range(10):
        a = ws.cell(2*i+16+22*j,3).value
        ws.cell(2*i+16+22*j,3).value = float(a.replace("'s':",""))
        b = ws.cell(2*i+16+22*j,4).value
        ws.cell(2*i+16+22*j,4).value = float(b.replace("'l':",""))
        c = ws.cell(2*i+16+22*j,2).value
        ws.cell(2*i+16+22*j,2).value = float(c[:1] +''+ c[12:]) #桁数によって変える必要がある．
        d = ws.cell(2*i+16+22*j,5).value
        ws.cell(2*i+16+22*j,5).value = d[:1] +''+ d[36:] #桁数によって変える必要がある．
        print(len(c),c)
        print(len(d),d)

for j in range(100): #状況によっては加える
    for i in range(10):
        d = ws.cell(2*i+16+22*j,5).value
        ws.cell(2*i+16+22*j,5).value = float(d.replace(':',''))

#ws.cell(row=8,column=2).value = lines[7].replace("Str =",'')

ws.cell(1,8).value = '協力的'
ws.cell(1,9).value = '懐疑的'
ws.cell(1,10).value = 'フリーライダー'
ws.cell(1,11).value = '嘘つき'
ws.cell(1,12).value = '回収個数'

for j in range(100):
    for i in range(10):
        ws.cell(i+2+11*j,8).value = ws.cell(2*i+16+22*j,2).value
        ws.cell(i+2+11*j,9).value = ws.cell(2*i+16+22*j,3).value
        ws.cell(i+2+11*j,10).value = ws.cell(2*i+16+22*j,4).value
        ws.cell(i+2+11*j,11).value = ws.cell(2*i+16+22*j,5).value
        ws.cell(i+2+11*j,12).value = ws.cell(2*i+15+22*j,6).value
    j = j+1

for i in range(100):
    ws.cell(11*i+12,8).value = '-----------------------'

for j in range(100):
    for i in range(10):
        ws.cell(j+2,i+15).value=ws.cell(i+2+11*j,8).value
        ws.cell(j+2,i+26).value=ws.cell(i+2+11*j,9).value
        ws.cell(j+2,i+37).value=ws.cell(i+2+11*j,10).value
        ws.cell(j+2,i+48).value=ws.cell(i+2+11*j,11).value
        ws.cell(j+2,i+59).value=ws.cell(i+2+11*j,12).value
    j = j+1

#for i in range(1,100):
    #ws.cell(i+1,15).value = '= AVERAGE(H:H11)'

#wb.save('繰り返しゲームga回収物.xlsx')
wb.save(bname)
