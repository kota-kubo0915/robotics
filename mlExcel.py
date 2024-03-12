import openpyxl
import re

tname = 'lml.txt'
bname = '回収物個数嘘つきml.xlsx'

file = open('lml.txt')
lines = file.readlines()

#wb = openpyxl.Workbook() #Book新規作成
wb = openpyxl.load_workbook(bname) #既存のBook読み込み
#wb.create_sheet(title='100 50set 500step')
ws = wb['100 10set 500step']

ws.cell(row=1,column=1).value = lines[0].split("\t")[0]
ws.cell(row=2,column=1).value = lines[1].split(":")[0]
ws.cell(row=2,column=2).value = float(lines[1].split(":")[1])

ws.cell(row=3,column=1).value = lines[2].split(":")[0]
ws.cell(row=3,column=2).value = float(lines[2].split(":")[1])

ws.cell(row=4,column=1).value = lines[3].split(":")[0]
ws.cell(row=4,column=2).value = float(lines[3].split(":")[1])

ws.cell(row=5,column=1).value = lines[4].split(":")[0]
ws.cell(row=5,column=2).value = float(lines[4].split(":")[1])


for i in range(1,100):
    ws.cell(row=5*i+1,column=1).value = lines[5*i].split("\t")[0]
    ws.cell(row=5*i+2,column=1).value = lines[5*i+1].split("\t")[0]

    ws.cell(row=5*i+3,column=1).value = lines[5*i+2].split("[")[1] and lines[5*i+2].split(",")[0]
    ws.cell(row=5*i+3,column=2).value = float(lines[5*i+2].split(",")[1])
    ws.cell(row=5*i+3,column=3).value = float(lines[5*i+2].split(",")[2])
    ws.cell(row=5*i+3,column=4).value = lines[5*i+2].split(",")[3] #and lines[5*i+2].split("]")[0]

    ws.cell(row=5*i+4,column=1).value = lines[5*i+3].split("[")[1] and lines[5*i+3].split(",")[0]
    ws.cell(row=5*i+4,column=2).value = float(lines[5*i+3].split(",")[1])
    ws.cell(row=5*i+4,column=3).value = float(lines[5*i+3].split(",")[2])
    ws.cell(row=5*i+4,column=4).value = lines[5*i+3].split(",")[3] #and lines[5*i+3].split("]")[0]

    ws.cell(row=5*i+5,column=1).value = lines[5*i+4].split("=")[0]
    ws.cell(row=5*i+5,column=2).value = float(lines[5*i+4].split("=")[1])

#ws.cell(row=8,column=5).value.split("]")[1]

#ws.cell(row=8,column=2).value = lines[7].replace("Str =",'')
for i in range(1,100):
    a = ws.cell(5*i+3,1).value
    ws.cell(5*i+3,1).value = float(a.replace('[',''))

    b = ws.cell(5*i+4,1).value
    ws.cell(5*i+4,1).value = float(b.replace('[',''))

    c = ws.cell(5*i+3,4).value
    ws.cell(5*i+3,4).value = float(c.replace(']',''))

    d = ws.cell(5*i+4,4).value
    ws.cell(5*i+4,4).value = float(d.replace(']',''))

ws.cell(1,6).value = '協力的'
ws.cell(1,7).value = '懐疑的'
ws.cell(1,8).value = 'フリーライダー'
ws.cell(1,9).value = '嘘つき'
ws.cell(1,10).value = '回収個数'

for i in range(1,100):
    ws.cell(i+1,6).value = ws.cell(5*i+3,1).value
    ws.cell(i+1,7).value = ws.cell(5*i+3,2).value
    ws.cell(i+1,8).value = ws.cell(5*i+3,3).value
    ws.cell(i+1,9).value = ws.cell(5*i+3,4).value
    ws.cell(i+1,10).value = ws.cell(5*i+5,2).value

ws.cell(1,13).value = '協力的'
ws.cell(1,14).value = '懐疑的'
ws.cell(1,15).value = 'フリーライダー'
ws.cell(1,16).value = '嘘つき'
ws.cell(1,17).value = '回収個数'

for i in range(100):
    ws.cell(i+1,13).value = ws.cell(row=5*i+4,column=1).value
    ws.cell(i+1,14).value = ws.cell(row=5*i+4,column=2).value
    ws.cell(i+1,15).value = ws.cell(row=5*i+4,column=3).value
    ws.cell(i+1,16).value = ws.cell(row=5*i+4,column=4).value
    ws.cell(i+1,17).value = ws.cell(5*i+5,2).value

#wb.save('回収物個数嘘つきml.xlsx')
wb.save(bname)
